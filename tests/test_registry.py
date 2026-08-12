import io
import os
import unittest

from openpyxl import load_workbook
from test_mandiri import MANDIRI_INPUT

import rekapimutasi
from rekapimutasi import (
    UnsupportedBankError,
    csv_bytes,
    flatten_statement,
    parse_text,
    xlsx_bytes,
)
from rekapimutasi.utils import parse_bca_balance, parse_bca_money, parse_idr


class RegistryTest(unittest.TestCase):
    def test_rejects_unknown(self):
        with self.assertRaises(UnsupportedBankError):
            parse_text("this is not a bank statement")

    def test_dispatch_mandiri(self):
        stmt = parse_text(MANDIRI_INPUT)
        self.assertEqual(stmt.bank, "MANDIRI")


BCA_REAL = "banks/bca/testdata/5771045236_JUL_2026.pdf"


@unittest.skipUnless(os.path.exists(BCA_REAL), "real BCA fixture not present")
class RegistryRealTest(unittest.TestCase):
    def test_dispatches_bca_pdf(self):
        stmt = parse_text(rekapimutasi.extract_pdf_text(BCA_REAL))
        self.assertEqual(stmt.bank, "BCA")


class ExportTest(unittest.TestCase):
    def setUp(self):
        self.stmt = parse_text(MANDIRI_INPUT)

    def test_csv_rows(self):
        rows = rekapimutasi.statement_csv_rows(self.stmt)
        self.assertEqual(rows[0][0], "Date")
        self.assertEqual(len(rows), 8)  # header + 7 transactions
        amount_col = rows[1][3]
        self.assertEqual(amount_col, "-5500")

    def test_csv_bytes_roundtrip(self):
        data = csv_bytes(self.stmt)
        text = data.decode("utf-8")
        self.assertIn("2023-09-05", text)
        self.assertIn("-5500", text)

    def test_xlsx_bytes(self):
        data = xlsx_bytes(self.stmt)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Transactions"]
        self.assertEqual(ws.cell(row=1, column=1).value, "Date")
        self.assertEqual(ws.cell(row=2, column=4).value, -5500)
        self.assertEqual(ws.cell(row=8, column=4).value, -12500)
        self.assertEqual(ws.max_row, 8)

    def test_flatten(self):
        flat = flatten_statement(self.stmt)
        self.assertEqual(flat["bank"], "MANDIRI")
        self.assertEqual(len(flat["rows"]), 7)
        self.assertEqual(flat["rows"][0][3], "-5500")


class MoneyParsingTest(unittest.TestCase):
    """Integer-only money parsing: no float truncation or rounding surprises."""

    def test_bca_money_credit(self):
        m = parse_bca_money("3,528,964.00", True)
        self.assertEqual(m.value, 3528964)

    def test_bca_money_suffix_wins(self):
        m = parse_bca_money("3,528,964.00 CR", False)
        self.assertEqual(m.value, 3528964)
        m = parse_bca_money("100,000.00 DB", True)
        self.assertEqual(m.value, -100000)

    def test_bca_money_debit(self):
        m = parse_bca_money("108400.00", False)
        self.assertEqual(m.value, -108400)

    def test_bca_money_large(self):
        m = parse_bca_money("2,085,794.13", True)
        self.assertEqual(m.value, 2085794)

    def test_balance(self):
        m = parse_bca_balance("25,441,156.00")
        self.assertEqual(m.value, 25441156)

    def test_idr_negative(self):
        m = parse_idr("-30.000")
        self.assertEqual(m.value, -30000)

    def test_idr_positive_with_decimal(self):
        m = parse_idr("61.832,04")
        self.assertEqual(m.value, 61832)

    def test_idr_plus(self):
        m = parse_idr("+221.861")
        self.assertEqual(m.value, 221861)


if __name__ == "__main__":
    unittest.main()
