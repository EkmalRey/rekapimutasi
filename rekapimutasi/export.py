import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font

_HEADERS = ["Date", "Time", "Type", "Amount", "Balance", "Source/Destination", "Details", "Notes", "Transaction ID"]

_AMOUNT_FMT = "+Rp#,##0;-Rp#,##0;Rp0"
_BALANCE_FMT = "Rp#,##0;-Rp#,##0;Rp0"


def _format_money(money):
    return "" if not money.currency else str(money.value)


def statement_csv_rows(stmt):
    rows = [_HEADERS]
    for pocket in stmt.pockets:
        for tx in pocket.transactions:
            rows.append(
                [
                    tx.date,
                    tx.time,
                    tx.mutation_type,
                    _format_money(tx.amount),
                    _format_money(tx.balance),
                    tx.source_destination,
                    tx.transaction_detail,
                    tx.notes,
                    tx.transaction_id,
                ]
            )
    return rows


def write_csv(stmt, out):
    writer = csv.writer(out)
    writer.writerows(statement_csv_rows(stmt))


def csv_bytes(stmt):
    buf = io.StringIO()
    write_csv(stmt, buf)
    return buf.getvalue().encode("utf-8")


def write_xlsx(stmt, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(_HEADERS)

    for pocket in stmt.pockets:
        if len(stmt.pockets) > 1 and pocket.name:
            cell = ws.cell(row=ws.max_row + 1, column=1, value=f"— {pocket.name} —")
            cell.font = Font(bold=True)
        for tx in pocket.transactions:
            ws.append(
                [
                    tx.date,
                    tx.time,
                    tx.mutation_type,
                    tx.amount.value,
                    tx.balance.value,
                    tx.source_destination,
                    tx.transaction_detail,
                    tx.notes,
                    tx.transaction_id,
                ]
            )

    for col, width in zip("ABCDEFGHI", [12, 8, 8, 18, 18, 40, 50, 30, 24]):
        ws.column_dimensions[col].width = width

    # Amount and Balance stay real numbers; Excel formats them as IDR via a
    # custom number format.
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = _AMOUNT_FMT
        ws.cell(row=row, column=5).number_format = _BALANCE_FMT

    wb.save(out)


def xlsx_bytes(stmt):
    buf = io.BytesIO()
    write_xlsx(stmt, buf)
    return buf.getvalue()


def compact_statement(stmt):
    parts = [f"{stmt.bank} account {stmt.account_no} ({stmt.account_name})"]
    for pocket in stmt.pockets:
        parts.append(f"  {pocket.name}: {len(pocket.transactions)} transactions")
    return "\n".join(parts)


def flatten_statement(stmt):
    """JSON shape consumed by the web frontend (rows flattened across pockets)."""
    rows = []
    for pocket in stmt.pockets:
        for tx in pocket.transactions:
            rows.append(
                [
                    tx.date,
                    tx.time,
                    tx.mutation_type,
                    _format_money(tx.amount),
                    _format_money(tx.balance),
                    tx.source_destination,
                    tx.transaction_detail,
                    tx.notes,
                    tx.transaction_id,
                ]
            )
    return {
        "bank": stmt.bank,
        "account_name": stmt.account_name,
        "account_no": stmt.account_no,
        "period": stmt.period,
        "currency": stmt.currency,
        "columns": _HEADERS,
        "rows": rows,
        "pockets": [p.name for p in stmt.pockets],
    }
